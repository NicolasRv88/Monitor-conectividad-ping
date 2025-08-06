import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label, Button
import threading
import platform
import time
from datetime import datetime
import winsound  # Para reproducir sonidos en Windows
import openpyxl  # Para trabajar con archivos Excel

# Asegurarse de que PyInstaller maneje correctamente los hilos
threading._main_thread = threading.current_thread()

def dentro_del_horario_restringido():
    """Verifica si la hora actual está dentro del rango restringido de 00:05 a 01:30."""
    hora_actual = datetime.now().time()
    inicio = datetime.strptime("00:05", "%H:%M").time()
    fin = datetime.strptime("01:30", "%H:%M").time()
    return inicio <= hora_actual <= fin

def mostrar_alerta(ip, host):
    """Muestra una ventana emergente con la alerta de un equipo sin comunicación, solo si está fuera del horario restringido."""
    if dentro_del_horario_restringido():
        return  # No mostrar alerta en el horario restringido

    ventana_alerta = Toplevel()
    ventana_alerta.title("Alerta de Conectividad")
    
    Label(ventana_alerta, text=f"Equipo sin comunicación:", font=("Arial", 12, "bold"), fg="red").pack(pady=5)
    Label(ventana_alerta, text=f"{host} ({ip})", font=("Arial", 10)).pack(pady=5)
    
    def cerrar_alerta():
        reconocer_alarmas()  # Detiene el sonido y limpia alertas
        ventana_alerta.destroy()  # Cierra la ventana

    Button(ventana_alerta, text="Aceptar", command=cerrar_alerta).pack(pady=10)

def reproducir_alerta():
    """Reproduce un sonido de alerta de Windows solo si está fuera del horario restringido."""
    if dentro_del_horario_restringido():
        return  # No reproducir sonido en el horario restringido

    try:
        winsound.PlaySound(r"C:\Windows\Media\Alarm06.wav", winsound.SND_FILENAME | winsound.SND_ASYNC)
    except Exception as e:
        print(f"Error al reproducir sonido: {str(e)}")
        
def reconocer_alarmas():
    """Reconoce las alarmas, detiene el sonido y oculta el botón de reconocimiento."""
    global alerta_activa
    alerta_activa = False
    winsound.PlaySound(None, winsound.SND_PURGE)  # Detiene el sonido de alerta
    boton_reconocer.pack_forget()  # Oculta el botón de reconocimiento

def realizar_ping(ip):
    """Realiza ping a una dirección IP y devuelve si responde o no."""
    try:
        sistema_operativo = platform.system().lower()
        if sistema_operativo == "windows":
            comando = ["ping", "-n", "1", "-w", "1000", ip]
        elif sistema_operativo in ["linux", "darwin"]:  # Darwin es para MacOS
            comando = ["ping", "-c", "1", "-W", "1", ip]
        else:
            raise ValueError(f"OS '{sistema_operativo}' no soportado para ping.")
        resultado = subprocess.run(
            comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
        )
        return resultado.returncode != 0  # True si no responde
    except Exception:
        return True  # Considerar como "No responde" si hay un error

def cargar_ips():
    """Carga las direcciones IP y nombres de host desde un archivo .txt."""
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo de IPs", filetypes=[("Archivos de texto", "*.txt")]
    )
    if archivo:
        try:
            with open(archivo, "r") as f:
                lineas = f.read().splitlines()
            ips_hosts.clear()
            for linea in lineas:
                partes = linea.split(maxsplit=1)  # Divide en IP y Host
                if len(partes) == 2 and validar_ip(partes[0]):
                    ips_hosts.append((partes[0], partes[1]))
            if not ips_hosts:
                messagebox.showerror("Error", "No se encontraron entradas válidas en el archivo.")
            else:
                actualizar_lista()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo: {str(e)}")

def validar_ip(ip):
    """Verifica si una dirección IP tiene un formato válido."""
    partes = ip.split(".")
    if len(partes) != 4:
        return False
    try:
        return all(0 <= int(parte) <= 255 for parte in partes)
    except ValueError:
        return False

def actualizar_lista():
    """Actualiza la lista de IPs y nombres de host en la interfaz."""
    texto_ips.config(state=tk.NORMAL)
    texto_ips.delete(1.0, tk.END)
    for ip, host in ips_hosts:
        texto_ips.insert(tk.END, f"{ip} {host}\n")
    texto_ips.config(state=tk.DISABLED)

def iniciar_consulta_continua():
    """Inicia el proceso de consulta continua de ping en un hilo."""
    global consulta_activa
    if consulta_activa:
        return  # Evita abrir otra ventana si la consulta ya está activa
    consulta_activa = True
    threading.Thread(target=consulta_continua, daemon=True).start()
    boton_rojo.config(state=tk.DISABLED)
    boton_verde.config(state=tk.NORMAL)
    parpadear_boton_verde()

def detener_consulta():
    """Detiene el proceso de consulta continua."""
    global consulta_activa
    consulta_activa = False
    boton_iniciar.config(state=tk.NORMAL)
    boton_detener.config(state=tk.DISABLED)
    boton_rojo.config(state=tk.NORMAL)
    boton_verde.config(state=tk.DISABLED)

def consulta_continua():
    """Realiza pings continuos a las direcciones IP cargadas."""
    global alerta_activa
    boton_iniciar.config(state=tk.DISABLED)
    boton_detener.config(state=tk.NORMAL)
    contador_no_responde = {ip: 0 for ip, _ in ips_hosts}  # Contador de no respuestas
    while consulta_activa:
        if ips_hosts:
            texto_resultados.config(state=tk.NORMAL)
            texto_resultados.delete(1.0, tk.END)
            for ip, host in ips_hosts:
                no_responde = realizar_ping(ip)
                if no_responde:
                    contador_no_responde[ip] += 1
                    if contador_no_responde[ip] >= 2:  # 2 intentos de 2 minutos = 4 minutos
                        texto_resultados.insert(tk.END, f"{ip} ({host}): No responde\n", "no_responde")
                        if ip not in historial_perdidas:
                            registro_perdida(ip, host)
                            alerta_activa = True
                            reproducir_alerta()
                            mostrar_alerta(ip, host)
                else:
                    contador_no_responde[ip] = 0  # Reiniciar contador si responde
            texto_resultados.config(state=tk.DISABLED)
        time.sleep(120)  # Consulta cada 2 minutos
    boton_iniciar.config(state=tk.NORMAL)
    boton_detener.config(state=tk.DISABLED)

def registro_perdida(ip, host):
    """Registra una pérdida de ping en el historial."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    historial_perdidas[ip] = f"{timestamp} - {ip} ({host}): No responde"
    guardar_en_archivo(f"{timestamp} - {ip} ({host}): No responde")
    actualizar_historial()

def actualizar_historial():
    """Actualiza el historial en la interfaz."""
    texto_historial.config(state=tk.NORMAL)
    texto_historial.delete(1.0, tk.END)
    for registro in sorted(historial_perdidas.values(), reverse=True):
        texto_historial.insert(tk.END, registro + "\n")
    texto_historial.config(state=tk.DISABLED)

def guardar_en_archivo(mensaje):
    """Guarda los registros en un archivo de texto y en un archivo Excel."""
    # Guardar en archivo de texto
    with open("historial_perdidas.txt", "a") as f:
        f.write(mensaje + "\n")
    
    # Guardar en archivo Excel
    try:
        archivo_excel = "historial_perdidas.xlsx"
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha y Hora", "IP", "Host", "Estado"])  # Encabezados

    # Agregar registro al Excel
    fecha_hora, resto = mensaje.split(" - ", 1)
    ip_host, estado = resto.rsplit(": ", 1)
    ip, host = ip_host.split(" (")
    host = host.rstrip(")")
    ws.append([fecha_hora, ip, host, estado])
    
    wb.save(archivo_excel)

def crear_archivo_excel():
    """Crea el archivo historial_perdidas.xlsx si no existe."""
    archivo_excel = "historial_perdidas.xlsx"
    try:
        openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha y Hora", "IP", "Host", "Estado"])  # Encabezados
        wb.save(archivo_excel)

def parpadear_boton_verde():
    """Hace parpadear el botón verde cuando el programa está en funcionamiento."""
    if consulta_activa:
        color_actual = boton_verde.cget("background")
        nuevo_color = "green" if color_actual != "green" else "light green"
        boton_verde.config(background=nuevo_color)
        root.after(500, parpadear_boton_verde)

# Interfaz gráfica
if __name__ == "__main__":
    # Asegurarse de que PyInstaller maneje correctamente los hilos
    threading._main_thread = threading.current_thread()
    
    crear_archivo_excel()  # Crear el archivo Excel si no existe
    root = tk.Tk()
    root.title("Ping a múltiples IPs - Consulta Continua")
    root.configure(bg="black")  # Fondo negro

    # Variables
    ips_hosts = []
    historial_perdidas = {}
    consulta_activa = False
    alerta_activa = False

    # Crear un frame para la barra superior
    frame_superior = tk.Frame(root, bg="black")
    frame_superior.pack(side=tk.TOP, fill=tk.X, pady=4)

    # Botones de estado en la parte superior derecha
    frame_botones_estado = tk.Frame(frame_superior, bg="black")
    frame_botones_estado.pack(side=tk.RIGHT, padx=8)

    boton_rojo = tk.Button(frame_botones_estado, bg="red", state=tk.NORMAL, width=2, height=1, bd=0, highlightthickness=0)
    boton_rojo.grid(row=0, column=0, padx=4)
    boton_rojo.config(relief=tk.FLAT, borderwidth=1, highlightbackground="white", highlightcolor="white")

    boton_verde = tk.Button(frame_botones_estado, bg="light green", state=tk.DISABLED, width=2, height=1, bd=0, highlightthickness=0)
    boton_verde.grid(row=0, column=1, padx=4)
    boton_verde.config(relief=tk.FLAT, borderwidth=1, highlightbackground="white", highlightcolor="white")

    boton_reconocer = tk.Button(frame_botones_estado, text="Reconocer Alarmas", command=reconocer_alarmas)
    boton_reconocer.grid(row=0, column=2, padx=4)

    # Marcos y widgets
    frame_carga = tk.Frame(root, bg="black")
    frame_carga.pack(pady=9)

    boton_cargar = tk.Button(frame_carga, text="Cargar IPs", command=cargar_ips)
    boton_cargar.pack(side=tk.LEFT, padx=5)

    boton_iniciar = tk.Button(frame_carga, text="Iniciar Consulta", command=iniciar_consulta_continua)
    boton_iniciar.pack(side=tk.LEFT, padx=5)

    boton_detener = tk.Button(frame_carga, text="Detener Consulta", command=detener_consulta, state=tk.DISABLED)
    boton_detener.pack(side=tk.LEFT, padx=5)

    # Listado de IP
    frame_ips = tk.Frame(root, bg="black")
    frame_ips.pack(pady=9)

    etiqueta_ips = tk.Label(frame_ips, text="Direcciones IP y nombres de host cargados:", fg="white", bg="black")
    etiqueta_ips.pack(anchor="w")

    scroll_ips = tk.Scrollbar(frame_ips)
    scroll_ips.pack(side=tk.RIGHT, fill=tk.Y)

    texto_ips = tk.Text(frame_ips, height=6, width=81, state=tk.DISABLED, yscrollcommand=scroll_ips.set, bg="black", fg="white")
    texto_ips.pack(side=tk.LEFT, fill=tk.BOTH)
    scroll_ips.config(command=texto_ips.yview)

    # Resultados (solo los que no responden)
    frame_resultados = tk.Frame(root, bg="black")
    frame_resultados.pack(pady=9)

    etiqueta_resultados = tk.Label(frame_resultados, text="Equipos que no responden al ping:", fg="white", bg="black")
    etiqueta_resultados.pack(anchor="w")

    scroll_resultados = tk.Scrollbar(frame_resultados)
    scroll_resultados.pack(side=tk.RIGHT, fill=tk.Y)

    texto_resultados = tk.Text(frame_resultados, height=13, width=81, state=tk.DISABLED, yscrollcommand=scroll_resultados.set, bg="black")
    texto_resultados.pack(side=tk.LEFT, fill=tk.BOTH)
    texto_resultados.tag_config("no_responde", foreground="red", font=("Arial",13, "bold")) # Texto rojo si no responde
    texto_resultados.tag_config("responde", foreground="green", font=("Arial", 10))  # Texto verde si responde
    scroll_resultados.config(command=texto_resultados.yview)

    # Marco histórico
    frame_historial = tk.Frame(root, bg="black")
    frame_historial.pack(pady=9)

    etiqueta_historial = tk.Label(frame_historial, text="Historial de pérdidas de ping:", fg="white", bg="black")
    etiqueta_historial.pack(anchor="w")

    scroll_historial = tk.Scrollbar(frame_historial)
    scroll_historial.pack(side=tk.RIGHT, fill=tk.Y)

    texto_historial = tk.Text(frame_historial, height=9, width=81, state=tk.DISABLED, yscrollcommand=scroll_historial.set, fg="white", bg="black")
    texto_historial.pack(side=tk.LEFT, fill=tk.BOTH)
    scroll_historial.config(command=texto_historial.yview)

    # Pie de página con texto
    frame_pie = tk.Frame(root, bg="black")
    frame_pie.pack(side=tk.BOTTOM, pady=8)

    etiqueta_pie = tk.Label(frame_pie, text="Desarrollado para Operadores_SCP", font=("Arial", 8), fg="green", bg="black")
    etiqueta_pie.pack(side=tk.LEFT, padx=4)

  # Interceptar el cierre de la ventana
    def on_closing():
        """Evita cerrar la ventana si la consulta está activa."""
        if consulta_activa:
            messagebox.showwarning("Aviso", "Debes detener la consulta antes de cerrar el programa.")
        else:
            root.destroy()  # Cierra la ventana solo si la consulta ha sido detenida

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # Ejecutar aplicación
    root.mainloop()